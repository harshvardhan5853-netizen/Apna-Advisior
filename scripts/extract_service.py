#!/usr/bin/env python3
"""
Apna Advisor — broker screenshot & PDF extraction service.

Pipeline:
  Images  → OpenCV preprocess → EasyOCR (+ Tesseract fallback) → heuristic parse
  PDFs    → pdfplumber tables/text → PyMuPDF text → scanned-page OCR fallback

Stdout: JSON { source, layout, engine, warnings, holdings, rawText }
"""

from __future__ import annotations

import argparse
import json
import math
import re
import sys
import uuid
from pathlib import Path
from typing import Any

# ── lazy singletons ──────────────────────────────────────────────────────────

_easyocr_reader = None
_tesseract_ok: bool | None = None


def uid(prefix: str = "h") -> str:
    return f"{prefix}-{uuid.uuid4().hex[:12]}"


def emit(payload: dict[str, Any]) -> None:
    sys.stdout.write(json.dumps(payload, ensure_ascii=False))
    sys.stdout.flush()


def emit_error(code: str, detail: str) -> None:
    emit({"error": code, "detail": detail, "holdings": [], "warnings": [detail]})


# ── broker detection ─────────────────────────────────────────────────────────

def detect_broker(text: str) -> str:
    t = text.lower()
    if re.search(r"\bgroww\b", t):
        return "groww"
    if re.search(r"\bzerodha\b|\bkite\b|\bconsole\b", t):
        return "zerodha"
    if re.search(r"\bangel one\b|\bangelone\b|\bangel broking\b", t):
        return "angelone"
    if re.search(r"\bupstox\b", t):
        return "upstox"
    if re.search(r"\bdhan\b", t):
        return "dhan"
    return "generic"


# ── number parsing ───────────────────────────────────────────────────────────

NUMBER_RE = re.compile(r"-?\(?\d[\d,]*\.?\d*\)?%?")
SIGNED_NUM_RE = re.compile(
    r"(?<![A-Za-z0-9])([+\-−])?\s*(?:₹|Rs\.?)?\s*(\d[\d,]*(?:\.\d+)?)\s*%?",
    re.IGNORECASE,
)
SHARES_AVG_RE = re.compile(
    r"(\d[\d,]*)\s*shares?\s*[^A-Za-z\d]*avg\.?\s*[₹Rs.]*\s*([\d,]+(?:\.\d+)?)",
    re.IGNORECASE,
)
HEADER_TOKENS = {
    "name", "names", "company", "instrument", "stock", "symbol", "qty", "quantity",
    "shares", "avg", "average", "ltp", "price", "market", "cmp", "inv", "invested",
    "investment", "amt", "amount", "current", "cur", "value", "val", "overall",
    "day", "days", "returns", "return", "gain", "loss", "g/l", "gl", "p&l", "pnl",
    "p/l", "holdings", "portfolio", "net", "chg", "change",
}


def parse_number_loose(raw: str) -> float:
    s = raw.strip().replace(",", "").replace("₹", "").replace("Rs.", "").replace("Rs", "")
    if s.endswith("%"):
        s = s[:-1]
    neg = s.startswith("(") and s.endswith(")")
    s = s.strip("()")
    try:
        n = float(s)
        return -n if neg else n
    except ValueError:
        return float("nan")


def extract_signed_numbers(text: str) -> list[float]:
    out: list[float] = []
    for m in SIGNED_NUM_RE.finditer(text):
        sign = -1 if m.group(1) in ("-", "−") else 1
        try:
            n = float(m.group(2).replace(",", ""))
            out.append(sign * n)
        except ValueError:
            continue
    return out


def correct_decimal_pair(qty: float, avg: float, invested: float) -> tuple[float, float]:
    if not (qty > 0 and avg > 0 and invested > 0):
        return avg, invested

    def near(a: float, b: float) -> bool:
        return abs(a / b - 1) < 0.08

    if near(qty * avg, invested):
        return avg, invested
    for div in (10, 100, 1000):
        if near(qty * (avg / div), invested):
            return avg / div, invested
    for div in (10, 100, 1000):
        if near(qty * avg, invested * div):
            return avg, invested * div
    return avg, invested


def score_row(qty: float, avg: float, ltp: float, invested: float, cur_val: float) -> float:
    score = 0.4
    if qty > 0:
        score += 0.08
    if avg > 0:
        score += 0.05
    if ltp > 0:
        score += 0.05
    if invested > 0:
        score += 0.05
    if cur_val > 0:
        score += 0.05

    def near(a: float, b: float) -> bool:
        return abs(a / b - 1) < 0.05

    if qty > 0 and avg > 0 and invested > 0 and near(qty * avg, invested):
        score += 0.16
    if qty > 0 and ltp > 0 and cur_val > 0 and near(qty * ltp, cur_val):
        score += 0.16
    return min(1.0, score)


def looks_like_header_row(text: str) -> bool:
    cleaned = re.sub(r"[^a-z0-9%/&\s]", " ", text.lower())
    tokens = [t for t in cleaned.split() if t]
    if not tokens:
        return False
    alpha = [t for t in tokens if re.match(r"^[a-z%/&]", t)]
    if not alpha:
        return False
    hits = sum(1 for t in alpha if t in HEADER_TOKENS)
    digit_count = len(re.findall(r"\d", cleaned))
    if hits >= 2 and digit_count <= 2:
        return True
    if hits >= 1 and len(alpha) <= 3 and digit_count <= 2:
        return True
    return False


def norm_text(s: str) -> str:
    return re.sub(r"\s{2,}", " ", s.replace("\u00a0", " ").replace("·", " ").replace("•", " ").strip())


def normalize_symbol(name: str) -> str:
    token = re.sub(r"[^A-Za-z0-9&-]", "", (name.split() or [""])[0]).upper()
    return token[:20] if token else "UNKNOWN"


def holding_from_parts(
    name: str,
    qty: float,
    avg: float,
    ltp: float,
    invested: float,
    cur_val: float,
    pnl: float | None,
    pnl_pct: float | None,
    source: str,
) -> dict[str, Any]:
    avg, invested = correct_decimal_pair(qty, avg, invested)
    ltp_fix = correct_decimal_pair(qty, ltp, cur_val)
    ltp, cur_val = ltp_fix[0], ltp_fix[1]
    if pnl is None:
        pnl = cur_val - invested
    if pnl_pct is None:
        pnl_pct = (pnl / invested) if invested > 0 else 0.0
    elif abs(pnl_pct) > 3:
        pnl_pct = pnl_pct / 100.0
    math_pnl = cur_val - invested
    if abs(math_pnl) > 0.5 and pnl != 0 and (pnl > 0) != (math_pnl > 0) and invested > 0:
        pnl = math_pnl
        pnl_pct = math_pnl / invested
    confidence = score_row(qty, avg, ltp, invested, cur_val)
    clean_name = re.sub(r"\s{2,}", " ", name.strip())
    return {
        "id": uid(),
        "stockName": clean_name,
        "symbol": normalize_symbol(clean_name),
        "exchange": "UNKNOWN",
        "quantity": int(qty) if qty == int(qty) else qty,
        "avgBuyPrice": round(avg, 4),
        "currentPrice": round(ltp, 4),
        "investedAmount": round(invested, 2),
        "currentValue": round(cur_val, 2),
        "pnl": round(pnl, 2),
        "pnlPercent": round(pnl_pct, 6),
        "confidence": round(confidence, 3),
        "needsReview": confidence < 0.8,
        "source": source,
    }


def extract_holdings_from_lines(lines: list[str], source: str) -> list[dict[str, Any]]:
    holdings: list[dict[str, Any]] = []
    for raw in lines:
        line = norm_text(raw.replace("\u00a0", " "))
        if not line:
            continue
        if re.match(r"^(stock|instrument|holdings|total|page|portfolio|summary)\b", line, re.I):
            continue
        nums = NUMBER_RE.findall(line)
        if len(nums) < 3:
            continue
        m = NUMBER_RE.search(line)
        if not m:
            continue
        first_idx = m.start()
        if first_idx <= 1:
            continue
        name_part = line[:first_idx].strip()
        if not re.search(r"[A-Za-z]", name_part):
            continue
        parsed = [parse_number_loose(n) for n in nums]
        parsed = [n for n in parsed if n == n]  # drop nan
        if len(parsed) < 3:
            continue
        qty, avg, cur, invested, cur_val, pnl = (parsed + [0] * 7)[:6]
        quantity = qty or 0
        avg_buy = avg or 0
        current_price = cur or 0
        invested_amt = invested if invested and invested > 0 else quantity * avg_buy
        current_value = cur_val if cur_val and cur_val > 0 else quantity * current_price
        pnl_val = pnl if pnl == pnl else current_value - invested_amt
        # Do NOT bind pnl_pct from parsed[6]: in broker screenshot layouts the 7th
        # number on the holding row is the Day's G/L amount (₹), not a percent, and
        # the real percentages sit on the next line. Let holding_from_parts recompute
        # pnl/invested, which is mathematically consistent and correct for every format.
        holdings.append(
            holding_from_parts(
                name_part, quantity, avg_buy, current_price,
                invested_amt, current_value, pnl_val, None, source,
            )
        )
    return holdings


def extract_card_layout(lines: list[str], source: str) -> list[dict[str, Any]]:
    holdings: list[dict[str, Any]] = []
    items = [norm_text(l) for l in lines if norm_text(l)]
    for i, text in enumerate(items):
        m = SHARES_AVG_RE.search(text)
        if not m:
            continue
        name = ""
        for j in range(i - 1, max(-1, i - 4), -1):
            t = items[j].strip()
            if not t or SHARES_AVG_RE.search(t) or looks_like_header_row(t):
                continue
            if re.match(r"^[+\-−₹\d(]", t):
                continue
            if not re.search(r"[A-Za-z]", t):
                continue
            name = re.sub(r"\s{2,}", " ", re.sub(r"[^\w &().-]", " ", t)).strip()
            break
        if not name:
            continue
        qty = parse_number_loose(m.group(1))
        avg = parse_number_loose(m.group(2))
        next_nums: list[float] = []
        for k in range(i + 1, min(len(items), i + 11)):
            t = items[k].strip()
            if not t:
                continue
            if SHARES_AVG_RE.search(t):
                break
            has_digits = bool(re.search(r"\d", t))
            has_words = bool(re.search(r"[A-Za-z]{3,}", t))
            if not has_digits and has_words and len(next_nums) >= 5:
                break
            next_nums.extend(extract_signed_numbers(t))
            if len(next_nums) >= 7:
                break
        if len(next_nums) >= 7:
            ltp, pnl, pnl_pct, cur_val, invested = next_nums[0], next_nums[3], next_nums[4], next_nums[5], next_nums[6]
        elif len(next_nums) == 6:
            ltp, pnl, pnl_pct, cur_val, invested = next_nums[0], next_nums[2], next_nums[3], next_nums[4], next_nums[5]
        elif len(next_nums) == 5:
            ltp, pnl, pnl_pct, cur_val, invested = next_nums[0], next_nums[1], next_nums[2], next_nums[3], next_nums[4]
        elif len(next_nums) >= 3:
            ltp = next_nums[0]
            cur_val = next_nums[-2]
            invested = next_nums[-1]
            pnl = cur_val - invested
            pnl_pct = (pnl / invested) if invested > 0 else 0
        else:
            continue
        holdings.append(holding_from_parts(name, qty, avg, ltp, invested, cur_val, pnl, pnl_pct, source))
    return holdings


def parse_text_to_holdings(text: str, source: str | None = None) -> tuple[list[dict[str, Any]], str]:
    broker = source or detect_broker(text)
    lines = [l.strip() for l in re.split(r"\r?\n", text) if l.strip()]
    card_hits = sum(1 for l in lines if SHARES_AVG_RE.search(l))
    layout = "card" if card_hits >= 2 or (broker == "groww" and card_hits >= 1) else "tabular"
    holdings = extract_card_layout(lines, broker) if layout == "card" else []
    if not holdings:
        holdings = extract_holdings_from_lines(lines, broker)
        layout = "tabular" if holdings else layout
    return holdings, layout


# ── image preprocessing (OpenCV) ─────────────────────────────────────────────

def preprocess_image(path: Path):
    import cv2
    import numpy as np

    img = cv2.imdecode(np.fromfile(str(path), dtype=np.uint8), cv2.IMREAD_COLOR)
    if img is None:
        raise ValueError(f"Could not decode image: {path}")
    h, w = img.shape[:2]
    target_w = 2000
    scale = max(1.0, min(3.0, target_w / max(1, w)))
    if scale > 1.01:
        img = cv2.resize(img, (int(w * scale), int(h * scale)), interpolation=cv2.INTER_CUBIC)
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
    gray = clahe.apply(gray)
    gray = cv2.fastNlMeansDenoising(gray, None, 10, 7, 21)
    _, thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    return thresh, img


def tesseract_available() -> bool:
    global _tesseract_ok
    if _tesseract_ok is not None:
        return _tesseract_ok
    try:
        import pytesseract
        # Windows winget install often isn't on PATH for child processes spawned by Node.
        if sys.platform == "win32":
            for candidate in (
                r"C:\Program Files\Tesseract-OCR\tesseract.exe",
                r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
            ):
                if Path(candidate).is_file():
                    pytesseract.pytesseract.tesseract_cmd = candidate
                    break
        pytesseract.get_tesseract_version()
        _tesseract_ok = True
    except Exception:
        _tesseract_ok = False
    return _tesseract_ok


def ocr_tesseract(img) -> tuple[str, str]:
    import pytesseract
    config = "--psm 6 -c preserve_interword_spaces=1"
    text = pytesseract.image_to_string(img, lang="eng", config=config)
    return text or "", "tesseract"


def ocr_easyocr(img) -> tuple[str, str]:
    global _easyocr_reader
    import easyocr
    import numpy as np

    if _easyocr_reader is None:
        _easyocr_reader = easyocr.Reader(["en"], gpu=False, verbose=False)
    arr = img if hasattr(img, "shape") else np.array(img)
    results = _easyocr_reader.readtext(arr, detail=0, paragraph=True)
    text = "\n".join(str(r).strip() for r in results if str(r).strip())
    return text, "easyocr"


def _debug_paddle(label: str, **fields: Any) -> None:
    """Emit structured debug info to stderr so it never corrupts stdout JSON."""
    import json as _json
    sys.stderr.write(f"[paddle-debug] {label}: {_json.dumps(fields, ensure_ascii=False)}\n")
    sys.stderr.flush()


def ocr_paddle(original_path: Path) -> tuple[str, str, list[str]]:
    """Primary screenshot OCR engine (PaddleOCR via smart_extract.py).

    Returns (flagged_text, engine, plain_rows) where ``plain_rows`` are
    pipe-free, space-joined cell strings ready for ``parse_text_to_holdings``.
    The flagged ``text`` is kept for human-facing rawText output.

    IMPORTANT: smart_extract.run_paddle_ocr decodes the file and performs its
    own 2× upscale internally — exactly as the standalone CLI does. We must
    pass it the ORIGINAL uploaded screenshot, NOT the binarized/CLAHE-preprocessed
    image, otherwise PaddleOCR sees a degraded image the standalone path never
    would (this was the integration regression).
    """
    import sys as _sys
    _sys.path.insert(0, str(Path(__file__).resolve().parent))
    from smart_extract import run_paddle_ocr

    _debug_paddle(
        "input",
        original_path=str(original_path),
        exists=str(original_path.is_file()),
    )

    result = run_paddle_ocr(str(original_path), verbose=False)

    flagged_text = result.get("text", "")
    plain_rows = result.get("plain_rows", [])
    raw_boxes = result.get("raw_boxes", [])
    _debug_paddle(
        "ocr",
        layout=result.get("layout"),
        n_flagged_text_chars=len(flagged_text),
        n_plain_rows=len(plain_rows),
        n_raw_boxes=len(raw_boxes),
        raw_boxes=raw_boxes[:50],
        plain_rows=plain_rows[:50],
        text=flagged_text[:2000],
        warnings=result.get("warnings", []),
    )
    return flagged_text, "paddleocr", plain_rows


def extract_image(path: Path) -> dict[str, Any]:
    warnings: list[str] = []
    try:
        processed, color = preprocess_image(path)
    except Exception as e:
        return {"error": "preprocess-failed", "detail": str(e), "holdings": [], "warnings": [str(e)]}

    try:
        import cv2
        import numpy as np
        _orig = cv2.imdecode(np.fromfile(str(path), dtype=np.uint8), cv2.IMREAD_COLOR)
        _debug_paddle(
            "image",
            original_path=str(path),
            dims=(_orig.shape[1], _orig.shape[0]) if _orig is not None else None,
            note="original screenshot passed directly to PaddleOCR (matches standalone smart_extract.py)",
        )
    except Exception:
        pass

    # ── Primary engine: PaddleOCR (smart_extract.py) ────────────────────────
    # NOTE: we pass the ORIGINAL screenshot path, not `processed` (the binarized
    # CLAHE image). smart_extract.run_paddle_ocr decodes + 2× upscales internally,
    # reproducing the standalone CLI exactly. Feeding it `processed` was the
    # integration regression (degraded OCR on a thresholded image).
    try:
        flagged_text, engine, plain_rows = ocr_paddle(path)
        # Feed CLEAN rows (no ⚠ markers) to the existing text heuristic.
        lines = plain_rows if plain_rows else [flagged_text]
        holdings, layout = parse_text_to_holdings("\n".join(lines))
        raw_text = flagged_text
        _debug_paddle(
            "parse",
            engine=engine,
            layout=layout,
            n_plain_rows=len(plain_rows),
            n_holdings=len(holdings),
            holdings=holdings,
        )
        if holdings:
            return _finalize_image_result(raw_text, engine, layout, holdings, warnings)
        # Primary engine produced no rows — fall through to legacy fallback.
        warnings.append("PaddleOCR ran but found no holdings rows — falling back to legacy OCR.")
    except Exception as e:
        warnings.append(f"PaddleOCR failed: {e}")

    # ── Legacy fallback: EasyOCR + Tesseract (kept, disabled as primary) ────
    candidates: list[tuple[str, str, list[dict[str, Any]]]] = []

    try:
        text, engine = ocr_easyocr(processed)
        holdings, layout = parse_text_to_holdings(text)
        candidates.append((engine, layout, holdings))
    except Exception as e:
        warnings.append(f"EasyOCR failed: {e}")

    if tesseract_available():
        try:
            text, engine = ocr_tesseract(processed)
            holdings, layout = parse_text_to_holdings(text)
            candidates.append((engine, layout, holdings))
        except Exception as e:
            warnings.append(f"Tesseract failed: {e}")
    else:
        warnings.append("Tesseract binary not found on PATH — install for fallback OCR.")

    if not candidates:
        return {
            "source": "generic",
            "layout": "unknown",
            "engine": "none",
            "warnings": warnings + ["All OCR engines failed."],
            "holdings": [],
            "rawText": "",
        }

    best = max(candidates, key=lambda c: (len(c[2]), sum(h.get("confidence", 0) for h in c[2])))
    engine, layout, holdings = best
    raw_text = ""
    try:
        raw_text, _ = ocr_easyocr(color) if engine == "easyocr" else ocr_tesseract(processed)
    except Exception:
        pass

    return _finalize_image_result(raw_text, engine, layout, holdings, warnings)


def _finalize_image_result(
    raw_text: str,
    engine: str,
    layout: str,
    holdings: list[dict[str, Any]],
    warnings: list[str],
) -> dict[str, Any]:
    broker = detect_broker(raw_text)
    if not holdings:
        warnings.append("OCR ran but no holdings rows detected — try Gemini or manual review.")
    return {
        "source": broker,
        "layout": layout,
        "engine": engine,
        "warnings": warnings,
        "holdings": holdings,
        "rawText": raw_text[:12000],
    }


# ── PDF extraction ───────────────────────────────────────────────────────────

def extract_pdf_text_lines(path: Path) -> tuple[list[str], list[str]]:
    lines: list[str] = []
    warnings: list[str] = []

    try:
        import pdfplumber
        with pdfplumber.open(str(path)) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables() or []
                for table in tables:
                    for row in table:
                        if not row:
                            continue
                        cells = [str(c or "").strip() for c in row]
                        line = "  ".join(c for c in cells if c)
                        if line:
                            lines.append(line)
                text = page.extract_text() or ""
                for ln in text.splitlines():
                    ln = ln.strip()
                    if ln:
                        lines.append(ln)
    except Exception as e:
        warnings.append(f"pdfplumber: {e}")

    if len(lines) < 3:
        try:
            import fitz  # PyMuPDF
            doc = fitz.open(str(path))
            try:
                pymupdf_lines: list[str] = []
                for page in doc:
                    text = page.get_text("text") or ""
                    for ln in text.splitlines():
                        ln = ln.strip()
                        if ln:
                            pymupdf_lines.append(ln)
                if len(pymupdf_lines) > len(lines):
                    lines = pymupdf_lines
            finally:
                doc.close()
        except Exception as e:
            warnings.append(f"PyMuPDF: {e}")

    return lines, warnings


def pdf_needs_ocr(path: Path) -> bool:
    try:
        import fitz
        doc = fitz.open(str(path))
        try:
            chars = sum(len((page.get_text() or "").strip()) for page in doc)
            return chars < 80 * max(1, doc.page_count)
        finally:
            doc.close()
    except Exception:
        return True


def ocr_pdf_pages(path: Path) -> tuple[str, list[str]]:
    import fitz
    warnings: list[str] = []
    texts: list[str] = []
    doc = fitz.open(str(path))
    try:
        for i, page in enumerate(doc):
            try:
                pix = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
                tmp = path.parent / f"_page_{i}.png"
                pix.save(str(tmp))
                result = extract_image(tmp)
                tmp.unlink(missing_ok=True)
                if result.get("rawText"):
                    texts.append(result["rawText"])
                elif result.get("holdings"):
                    texts.extend(
                        f"{h.get('stockName','')} {h.get('quantity','')} {h.get('avgBuyPrice','')} {h.get('investedAmount','')}"
                        for h in result["holdings"]
                    )
            except Exception as e:
                warnings.append(f"Page {i + 1} OCR failed: {e}")
    finally:
        doc.close()
    return "\n".join(texts), warnings


def extract_pdf(path: Path) -> dict[str, Any]:
    warnings: list[str] = []
    lines, w = extract_pdf_text_lines(path)
    warnings.extend(w)

    raw_text = "\n".join(lines)
    holdings, layout = parse_text_to_holdings(raw_text)
    engine = "pdfplumber"

    if not holdings and pdf_needs_ocr(path):
        warnings.append("PDF appears scanned — running page OCR.")
        ocr_text, ocr_w = ocr_pdf_pages(path)
        warnings.extend(ocr_w)
        if ocr_text:
            raw_text = ocr_text
            holdings, layout = parse_text_to_holdings(ocr_text)
            engine = "pdf-ocr"

    broker = detect_broker(raw_text)
    if not holdings:
        warnings.append("PDF parsed but no holdings rows found — try a screenshot or Gemini.")

    return {
        "source": broker,
        "layout": layout,
        "engine": engine,
        "warnings": warnings,
        "holdings": holdings,
        "rawText": raw_text[:12000],
    }


# ── entrypoint ───────────────────────────────────────────────────────────────

OLE2_HEADER = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


def check_protection(path: Path) -> dict[str, Any]:
    """Classify a file as normal / password-protected / corrupted / unsupported."""
    ext = path.suffix.lower()

    # ── PDF via PyMuPDF ──────────────────────────────────────────────────
    if ext == ".pdf":
        try:
            import fitz
        except ImportError:
            return {"unsupported": True, "format": "pdf",
                    "detail": "This file format is not supported (missing PyMuPDF)."}
        try:
            doc = fitz.open(str(path))
            encrypted = doc.is_encrypted
            doc.close()
            if encrypted:
                return {"passwordProtected": True, "format": "pdf"}
            return {"passwordProtected": False, "format": "pdf"}
        except Exception as e:
            exc_name = type(e).__name__
            if exc_name in ("FileDataError", "EmptyFileError"):
                return {"corrupted": True, "format": "pdf",
                        "detail": "This file appears to be corrupted or unreadable."}
            return {"unsupported": True, "format": "pdf",
                    "detail": "This file format is not supported."}

    # ── Excel via zipfile / OLE2 marker ──────────────────────────────────
    if ext in (".xlsx", ".xls", ".xlsm"):
        try:
            import zipfile
            with zipfile.ZipFile(path) as z:
                namelist_lower = [name.lower() for name in z.namelist()]
                # Modern OOXML Standard Encryption embeds encrypted content
                # markers inside a valid ZIP archive.
                if "encryptedpackage" in namelist_lower or "encryptioninfo" in namelist_lower:
                    return {"passwordProtected": True, "format": "xlsx"}
                return {"passwordProtected": False, "format": "xlsx"}
        except zipfile.BadZipFile:
            try:
                with open(path, "rb") as f:
                    header = f.read(8)
                if header == OLE2_HEADER:
                    # OLE2 (CFB) container — .xls files use this natively;
                    # .xlsx in this format indicates encryption.
                    return {"passwordProtected": True, "format": "xlsx"}
                return {"corrupted": True, "format": "xlsx",
                        "detail": "This file appears to be corrupted or unreadable."}
            except Exception:
                return {"corrupted": True, "format": "xlsx",
                        "detail": "This file appears to be corrupted or unreadable."}
        except Exception:
            return {"unsupported": True, "format": "xlsx",
                    "detail": "This file format is not supported."}

    # ── Other formats are never password-protected ───────────────────────
    return {"passwordProtected": False, "format": "unknown"}


def validate_password(path: Path, password: str) -> dict[str, Any]:
    """Try opening a file with the given password."""
    ext = path.suffix.lower()

    if ext == ".pdf":
        try:
            import fitz
            doc = fitz.open(str(path))
            if doc.is_encrypted:
                doc.close()
                doc = fitz.open(str(path))
                doc.authenticate(password)
            if not doc.is_encrypted:
                doc.close()
                return {"passwordOk": True, "format": "pdf"}
            doc.close()
            return {"passwordOk": False, "format": "pdf"}
        except Exception as e:
            exc_name = type(e).__name__
            if exc_name in ("FileDataError", "EmptyFileError"):
                return {"passwordOk": False, "format": "pdf",
                        "detail": "This file appears to be corrupted or unreadable."}
            return {"passwordOk": False, "format": "pdf"}

    if ext in (".xlsx", ".xls", ".xlsm"):
        try:
            import msoffcrypto
            import tempfile
            with open(path, "rb") as f:
                office = msoffcrypto.OfficeFile(f)
                office.load_key(password=password, verify_password=True)
                # Fully decrypt to temp to confirm the key is valid
                with tempfile.TemporaryFile() as decrypted:
                    office.decrypt(decrypted)
            return {"passwordOk": True, "format": "xlsx"}
        except Exception as e:
            exc_name = type(e).__name__
            # msoffcrypto raises InvalidKeyError on wrong password
            if exc_name == "InvalidKeyError" or "invalid key" in str(e).lower():
                return {"passwordOk": False, "format": "xlsx"}
            return {"passwordOk": False, "format": "xlsx",
                    "detail": f"XLSX validation failed: {e}"}

    return {"passwordOk": False, "format": "unknown"}


# ── XLSX extraction (msoffcrypto + openpyxl) ─────────────────────────────


def extract_xlsx(path: Path, password: str) -> dict[str, Any]:
    """Decrypt (if needed) and extract holdings from an Excel file."""
    warnings: list[str] = []
    import tempfile
    import os

    # ── Step 1: decrypt if password-protected ─────────────────────────────
    if password:
        try:
            import msoffcrypto
            with open(path, "rb") as f_in:
                office = msoffcrypto.OfficeFile(f_in)
                office.load_key(password=password, verify_password=True)
                tmp_fd, tmp_path_str = tempfile.mkstemp(suffix=".xlsx")
                os.close(tmp_fd)
                with open(tmp_path_str, "wb") as f_out:
                    office.decrypt(f_out)
                working_path = Path(tmp_path_str)
                _cleanup_tmp = tmp_path_str
        except Exception as e:
            exc_name = type(e).__name__
            if exc_name == "InvalidKeyError" or "invalid key" in str(e).lower():
                return {"error": "wrong-password", "detail": "That password didn't work. Passwords are case-sensitive."}
            try:
                check = check_protection(path)
                if check.get("corrupted"):
                    return {"error": "corrupted", "detail": "This file appears to be corrupted or unreadable."}
                if check.get("unsupported"):
                    return {"error": "unsupported", "detail": "This file format is not supported."}
                if check.get("passwordProtected"):
                    return {"error": "wrong-password", "detail": "That password didn't work. Passwords are case-sensitive."}
                working_path = path
                _cleanup_tmp = None
            except Exception:
                return {"error": "decrypt-failed", "detail": f"Could not decrypt XLSX: {e}"}
    else:
        check = check_protection(path)
        if check.get("corrupted"):
            return {"error": "corrupted", "detail": "This file appears to be corrupted or unreadable."}
        if check.get("unsupported"):
            return {"error": "unsupported", "detail": "This file format is not supported."}
        if check.get("passwordProtected"):
            return {"error": "wrong-password", "detail": "This file is password-protected but no password was provided."}
        working_path = path
        _cleanup_tmp = None

    # ── Step 2: read with openpyxl ───────────────────────────────────────
    try:
        import openpyxl
        wb = openpyxl.load_workbook(working_path, read_only=False, data_only=True)
    except Exception as e:
        if _cleanup_tmp:
            os.unlink(_cleanup_tmp)
        return {"error": "xlsx-read-failed", "detail": f"Could not read XLSX: {e}"}

    if _cleanup_tmp:
        try:
            os.unlink(_cleanup_tmp)
        except OSError:
            pass

    # ── Step 3: column-based holdings extraction ──────────────────────────
    # Header synonyms (mirrors column-map.ts on the client side).
    HEADER_SYNONYMS = {
        "stockName": ["stock name", "stock", "instrument", "company", "company name",
                       "security", "scrip", "scrip name", "name"],
        "symbol": ["symbol", "ticker", "trading symbol", "tradingsymbol",
                    "nse symbol", "bse symbol", "isin"],
        "exchange": ["exchange", "exch", "segment"],
        "quantity": ["qty", "quantity", "qty.", "holdings", "shares",
                     "no. of shares", "no of shares", "total quantity"],
        "avgBuyPrice": ["avg price", "average price", "avg. price", "avg cost",
                        "avg buy price", "average buy price", "buy avg", "buy price",
                        "cost price", "avg trading price"],
        "currentPrice": ["ltp", "last traded price", "current price", "cmp",
                         "market price", "close", "closing price"],
        "investedAmount": ["invested", "invested amount", "investment",
                           "cost value", "buy value", "amount invested", "invested value"],
        "currentValue": ["current value", "market value", "cur. val", "cur value", "present value"],
        "pnl": ["p&l", "pnl", "profit/loss", "gain/loss", "unrealized p&l", "net p&l", "overall gain/loss"],
        "pnlPercent": ["p&l %", "pnl %", "net chg.", "% change", "returns %",
                       "return %", "chg %", "overall gain/loss(%)"],
    }

    def norm_header(text: str) -> str:
        return text.lower().replace("_", " ").replace(".", " ").strip()

    def detect_columns(headers: list[str]) -> dict[str, int]:
        """Return a dict like {stockName: 0, quantity: 4, ...} with -1 for missing."""
        normed = [norm_header(h) for h in headers]
        result: dict[str, int] = {}
        for field, syns in HEADER_SYNONYMS.items():
            found = -1
            for i, h in enumerate(normed):
                if h and any(s == h or s in h.split() for s in syns):
                    found = i
                    break
            result[field] = found
        return result

    def safe_float(v) -> float:
        if v is None:
            return float("nan")
        try:
            return float(str(v).replace(",", "").strip())
        except (ValueError, TypeError):
            return float("nan")

    def parse_sheet(ws) -> list[dict]:
        """Extract holdings from a single worksheet using column detection."""
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        # Find the first row that looks like a holdings header.
        # Must have at least 2 known columns AND include stockName or symbol
        # (summary/overview sheets may have invested/market value columns but
        # no per-stock data, so they must be excluded).
        header_idx = -1
        header_map: dict[str, int] = {}
        for i, row in enumerate(rows):
            if not row:
                continue
            cell_strs = [str(c).strip() if c is not None else "" for c in row]
            m = detect_columns(cell_strs)
            known = sum(1 for v in m.values() if v >= 0)
            has_stock = m.get("stockName", -1) >= 0 or m.get("symbol", -1) >= 0
            if known >= 2 and has_stock:
                header_idx = i
                header_map = m
                break

        if header_idx < 0:
            return []

        holdings: list[dict] = []
        for row in rows[header_idx + 1:]:
            if not row:
                continue
            row_strs = [str(c).strip() if c is not None else "" for c in row]

            # Skip empty rows and total rows.
            if not any(c for c in row_strs):
                continue
            first = row_strs[0]
            if first.lower() in ("total", "grand total", "subtotal", "sum"):
                break

            # Extract key fields.
            stock_name = row_strs[header_map["stockName"]] if header_map.get("stockName", -1) >= 0 else ""
            qty = safe_float(row[header_map["quantity"]]) if header_map.get("quantity", -1) >= 0 else float("nan")
            avg = safe_float(row[header_map["avgBuyPrice"]]) if header_map.get("avgBuyPrice", -1) >= 0 else float("nan")
            ltp = safe_float(row[header_map["currentPrice"]]) if header_map.get("currentPrice", -1) >= 0 else float("nan")
            inv = safe_float(row[header_map["investedAmount"]]) if header_map.get("investedAmount", -1) >= 0 else float("nan")
            cur_val = safe_float(row[header_map["currentValue"]]) if header_map.get("currentValue", -1) >= 0 else float("nan")
            pnl = safe_float(row[header_map["pnl"]]) if header_map.get("pnl", -1) >= 0 else float("nan")
            pnl_pct = safe_float(row[header_map["pnlPercent"]]) if header_map.get("pnlPercent", -1) >= 0 else float("nan")

            # Must have stock name and at least one numeric to be a real holding.
            if not stock_name or (not any(math.isfinite(n) for n in (qty, avg, ltp, inv, cur_val))):
                continue

            # Back-fill missing values.
            if not math.isfinite(inv) and math.isfinite(qty) and math.isfinite(avg):
                inv = qty * avg
            if not math.isfinite(cur_val) and math.isfinite(qty) and math.isfinite(ltp):
                cur_val = qty * ltp
            if not math.isfinite(pnl) and math.isfinite(cur_val) and math.isfinite(inv):
                pnl = cur_val - inv
            if not math.isfinite(pnl_pct) and math.isfinite(pnl) and math.isfinite(inv) and inv > 0:
                pnl_pct = pnl / inv

            # Confidence.
            essentials_present = sum(1 for n in (qty, avg, ltp) if math.isfinite(n))
            confidence = 1.0 if essentials_present == 3 else (0.8 if essentials_present == 2 else 0.5)

            import re
            def normalize_symbol(name: str) -> str:
                token = re.sub(r"[^A-Za-z0-9&-]", "", (name.split() or [""])[0]).upper()
                return token[:20] if token else "UNKNOWN"

            clean_name = re.sub(r"\s{2,}", " ", stock_name.strip())
            holdings.append({
                "stockName": clean_name,
                "symbol": normalize_symbol(clean_name),
                "exchange": "UNKNOWN",
                "quantity": int(qty) if math.isfinite(qty) and qty == int(qty) else (qty if math.isfinite(qty) else 0),
                "avgBuyPrice": round(avg, 4) if math.isfinite(avg) else 0,
                "currentPrice": round(ltp, 4) if math.isfinite(ltp) else 0,
                "investedAmount": round(inv, 2) if math.isfinite(inv) else 0,
                "currentValue": round(cur_val, 2) if math.isfinite(cur_val) else 0,
                "pnl": round(pnl, 2) if math.isfinite(pnl) else 0,
                "pnlPercent": round(pnl_pct, 6) if math.isfinite(pnl_pct) else 0,
                "confidence": round(confidence, 3),
                "needsReview": confidence < 0.8,
                "source": "generic",
            })

        return holdings

    all_holdings: list[dict] = []
    all_rows_text: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_holdings = parse_sheet(ws)
        all_holdings.extend(sheet_holdings)
        # Also collect text rows for rawText / Gemini fallback.
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            line = "  ".join(c for c in cells if c)
            if line:
                all_rows_text.append(line)
    wb.close()

    raw_text = "\n".join(all_rows_text)
    broker = detect_broker(raw_text)

    if not all_holdings:
        # Fall back to the text heuristic if column detection found nothing.
        holdings, layout = parse_text_to_holdings(raw_text, broker)
        if holdings:
            warnings.append("XLSX parsed via text heuristic (column detection found no recognizable headers).")
        else:
            warnings.append("XLSX parsed but no holdings rows found — try Gemini or manual entry.")
        return {
            "source": broker,
            "layout": layout,
            "engine": "xlsx",
            "warnings": warnings,
            "holdings": holdings,
            "rawText": raw_text[:12000],
        }

    return {
        "source": broker,
        "layout": "tabular",
        "engine": "xlsx",
        "warnings": warnings,
        "holdings": all_holdings,
        "rawText": raw_text[:12000],
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Apna Advisor file extraction")
    parser.add_argument("file", type=str, help="Path to file")
    parser.add_argument("--kind", choices=["auto", "image", "pdf", "xlsx"], default="auto")
    parser.add_argument("--check-protection", action="store_true", help="Only check if file is password-protected")
    parser.add_argument("--password-stdin", action="store_true", help="Read password from stdin (secure)")
    parser.add_argument("--extract-with-password", action="store_true", help="Extract XLSX with password from stdin")
    args = parser.parse_args()

    path = Path(args.file)
    if not path.is_file():
        emit_error("file-not-found", f"File not found: {path}")
        return 1

    ext = path.suffix.lower()

    # ── Protection check mode ──
    if args.check_protection:
        result = check_protection(path)
        emit(result)
        return 0

    # ── Password validation mode (password via stdin) ──
    if args.password_stdin:
        password = sys.stdin.readline().strip()
        result = validate_password(path, password)
        password = None  # Clear from memory immediately
        emit(result)
        return 0

    # ── XLSX extraction with password via stdin ──
    if args.extract_with_password:
        password = sys.stdin.readline().strip()
        result = extract_xlsx(path, password)
        password = None  # Clear from memory immediately
        if "error" in result and result["error"] not in ("wrong-password",):
            emit(result)
            return 1
        emit(result)
        return 0

    # ── Normal extraction mode ──
    kind = args.kind
    if kind == "auto":
        if ext in (".xlsx", ".xls", ".xlsm"):
            kind = "xlsx"
        elif ext == ".pdf":
            kind = "pdf"
        else:
            kind = "image"

    try:
        if kind == "xlsx":
            result = extract_xlsx(path, "")
        elif kind == "pdf":
            result = extract_pdf(path)
        else:
            result = extract_image(path)
        if "error" in result and not result.get("holdings"):
            emit(result)
            return 1
        emit(result)
        return 0
    except Exception as e:
        emit_error("extract-failed", str(e))
        return 1


if __name__ == "__main__":
    sys.exit(main())
